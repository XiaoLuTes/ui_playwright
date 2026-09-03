from time import sleep
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from utils.logger import logger
from utils.element_locator import ElementLocator
from config.settings import settings
import allure
import time
import os
from pathlib import Path
from playwright.sync_api import expect


class BasePage:
    """页面基类"""

    def __init__(self, page_name, page):
        self.page = page  # 页面对象
        self.page_name = page_name  # 页面名
        self.element_locator = ElementLocator()  # 元素定位器实例
        self.locators = self.element_locator.load_locators()  # 获取所有页面元素
        self.settings = settings  # 配置项
        self.wait_timeout = self.settings.IMPLICIT_WAIT
        self._db_utils = None
        self.page_manager = None
        self.page.set_default_timeout(self.wait_timeout)

    def set_page_manager(self, page_manager):
        # 设置页面管理器
        self.page_manager = page_manager

    def get_element_locator(self, element_name):
        # 获取元素定位器
        page_name_str = self.find_element_page(element_name)
        return self.element_locator.get_locator(page_name_str, element_name)

    def set_db_utils(self, db_utils):
        self._db_utils = db_utils
        logger.debug(f"页面 '{self.page_name}' 已注入数据库连接")

    @allure.step("打开页面: {url}")
    def open(self, url):
        logger.info(f"打开页面: {url}")
        self.page.goto(url, wait_until="domcontentloaded")
        logger.info(f"页面打开成功: {url}")

    def find_element(self, element_name, screenshot_on_error=True, timeout=None):
        by, value = self.get_element_locator(element_name)
        is_hidden = element_name.startswith("hidden_")
        by = by.lower()
        if by == "id":
            loc = self.page.locator(f"#{value}")
        elif by == "xpath":
            loc = self.page.locator(f"xpath={value}")
        elif by == "css":
            loc = self.page.locator(value)
        elif by == "name":
            loc = self.page.locator(f"[name='{value}']")
        else:
            loc = self.page.locator(value)
        loc = loc.first  # 如果找到多个元素,则返回第一个

        try:
            if is_hidden:
                loc.wait_for(state="attached", timeout=timeout)
            else:
                loc.wait_for(state="visible", timeout=timeout)
            return loc
        except PlaywrightTimeoutError:
            error_msg = f"元素查找超时: {element_name}"
            logger.error(error_msg)
            if screenshot_on_error:
                self.take_screenshot(f"元素查找超时-{element_name}")
            raise PlaywrightTimeoutError(error_msg)
        except Exception as e:
            error_msg = f"元素{element_name}查找失败: {e}"
            logger.error(error_msg)
            if screenshot_on_error:
                self.take_screenshot(f"元素查找失败-{element_name}")
            raise e

    @allure.step("对元素【{element_name}】输入文本: {text}")
    def input_text(self, element_name, text, clear_first=True):
        logger.info(f"元素【{element_name}】输入: {text}")
        loc = self.find_element(element_name)
        if clear_first:
            loc.fill("")
        loc.fill(str(text))

    @allure.step("点击元素: {element_name}")
    def element_click(self, element_name, retry_times=3):
        """点击元素：真实点击；失败时检查存在性+可点击状态，就绪后重试"""
        loc = self.find_element(element_name)
        loc.scroll_into_view_if_needed()

        for attempt in range(1, retry_times + 1):
            try:
                loc.click()
                logger.info(f"点击成功: {element_name}")
                return
            except TimeoutError as e:
                # 预期内：元素暂时不可点击（不可见/不稳定/被遮挡），走重试流程
                logger.warning(f"第{attempt}次点击超时(不可点击): {element_name} => {e}")
            except Exception as e:
                # 非超时异常（定位器语法错、页面崩溃等）
                logger.error(f"点击异常(非超时): {element_name} => {e}")
                raise
            if attempt == retry_times:
                break
            # 元素存在性检查
            if not self.is_element_present(element_name, screenshot_on_error=False, timeout=3000):
                logger.error(f"点击失败后元素已不存在: {element_name}")
                break
            # 等待可点击状态（预期异常：等待超时/断言失败）
            try:
                expect(loc).to_be_visible(timeout=3000)
                expect(loc).to_be_enabled(timeout=3000)
            except (AssertionError, TimeoutError) as wait_err:
                logger.warning(f"元素未能就绪: {element_name} => {wait_err}")
            time.sleep(1)

        error_msg = f"点击失败(重试{retry_times}次): {element_name}"
        self.take_screenshot(f"点击失败-{element_name}")
        raise Exception(error_msg)

    def _get_text(self, element_name, timeout=None, screenshot_on_error=False):
        # 内部封装去掉装饰器,轮询时不需要触发allure.step,用于等待元素文本变更后比对
        loc = self.find_element(element_name, timeout=timeout, screenshot_on_error=screenshot_on_error)
        return loc.text_content().strip()

    @allure.step("获取元素文本: {element_name}")
    def get_text(self, element_name, timeout=None, screenshot_on_error=False):
        # 供executor-check_text、save_text使用
        for attempt in range(1, 4):  # 防止元素已出现，但前端未取值
            value = self._get_text(element_name, timeout=timeout, screenshot_on_error=screenshot_on_error)
            if value:
                return value
            logger.info(f"元素:{element_name}文本为空,等待1s后重试")
            time.sleep(1)
        logger.warning(f"元素:{element_name}文本仍为空,已重试三次")
        return ""

    def _get_element_value(self, element_name, timeout=None, screenshot_on_error=False):
        # 内部封装去掉装饰器,轮询时不需要触发allure.step,用于等待元素值变更后比对
        loc = self.find_element(element_name, timeout=timeout, screenshot_on_error=screenshot_on_error)
        return loc.input_value().strip()

    @allure.step("获取元素值: {element_name}")
    def get_element_value(self, element_name, timeout=None, screenshot_on_error=False):
        # 专供executor-check_value、save_value使用
        for attempt in range(1, 4):  # 防止元素已出现，但前端未取值
            value = self._get_element_value(element_name, timeout=timeout, screenshot_on_error=screenshot_on_error)
            if value:
                return value
            logger.info(f"元素:{element_name}值为空,等待1s后重试")
            time.sleep(1)
        logger.warning(f"元素:{element_name}值仍为空,已重试三次")
        return ""

    def is_element_present(self, element_name, screenshot_on_error=False, timeout=None):
        # 校验元素是否存在
        try:
            self.find_element(element_name, screenshot_on_error=screenshot_on_error, timeout=timeout)
            return True
        except PlaywrightTimeoutError:
            return False

    @allure.step("页面截图")
    def take_screenshot(self, name="screenshot"):
        screenshot_dir = self.settings.SCREENSHOT_PATH
        os.makedirs(screenshot_dir, exist_ok=True)
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"{name}_{timestamp}.png"
        filepath = os.path.join(screenshot_dir, filename)

        self.page.screenshot(path=filepath, full_page=True)
        logger.info(f"截图已保存: {filepath}")

        allure.attach.file(
            filepath,
            name=name,
            attachment_type=allure.attachment_type.PNG
        )

    def find_element_page(self, element_name):
        for page_name, elements in self.locators.items():
            if element_name in elements:
                return page_name
        return None

    @allure.step("键盘向下按键 {data} 次")
    def keyboard_down(self, data):
        times = int(data)
        for _ in range(times):
            self.page.keyboard.press("ArrowDown")
        logger.info(f"向下按键 {times} 次")

    @allure.step("键盘向上按键 {data} 次")
    def keyboard_up(self, data):
        times = int(data)
        for _ in range(times):
            self.page.keyboard.press("ArrowUp")
        logger.info(f"向上按键 {times} 次")

    @allure.step("按下 Enter 键")
    def keyboard_enter(self):
        self.page.keyboard.press("Enter")
        logger.info("按下 Enter 键")

    @allure.step("下载文件")
    def download_file(self, element_name, save_filename):
        """点击元素触发浏览器下载，保存到项目 reports/downloads 目录
        :param element_name: 触发下载的元素（下载按钮）
        :param save_filename: 保存的文件名（如 template.xlsx）
        :return: 保存后的完整路径
        """
        # 目录定位：相对路径统一基于项目根
        save_dir = self.settings.DOWNLOAD_DIR
        if not os.path.isabs(save_dir):  # 检查目录是否是绝对路径
            save_dir = os.path.join(self.settings.PROJECT_ROOT, save_dir)
        os.makedirs(save_dir, exist_ok=True)
        # Playwright 下载监听
        with self.page.expect_download() as dl_info:
            self.element_click(element_name)  # 触发下载（带防遮挡+重试）
        download = dl_info.value  # 拿到下载事件
        # 保存到本地（save_as 只能调用一次）
        save_path = os.path.join(save_dir, save_filename)
        download.save_as(save_path)
        logger.info(f"文件已下载: {save_path}")
        return save_path

    def edit_excel(self, file_path, cells):
        """填写 Excel 单元格并保存（覆盖原文件，供后续上传）
        :param file_path: 下载的模板文件路径
        :param cells: 填写配置 {"单元格": "值"}，如 {"A1": "张三", "B2": "20260903"}
        :return: 保存后的文件路径（原路径，内容已更新）
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        from openpyxl import load_workbook
        wb = load_workbook(file_path)  # 打开已有 Excel（保留原格式）
        ws = wb.active  # 默认第一个 sheet
        for cell, value in cells.items():
            ws[cell] = value  # 逐个单元格写入
        wb.save(file_path)  # 覆盖保存（路径不变）
        wb.close()  # 关闭释放文件锁（Windows 上传前必须）
        logger.info(f"Excel已填写: {file_path} cells={cells}")
        return file_path

    @allure.step("上传文件: {data}")
    def upload_file(self, element_name, data):
        if not Path(data).exists():
            raise FileNotFoundError(f"文件不存在: {data}")

        abs_path = str(Path(data).resolve())
        loc = self.find_element(element_name)
        loc.set_input_files(abs_path)
        logger.info(f"文件上传成功: {abs_path}")
        time.sleep(3)

    @allure.step("等待元素【{element_name}】的【{real_action}】等于【{expected_value}】")
    def wait_for_element_value(self, element_name, real_action, expected_value):
        timeout = self.settings.EXPLICIT_WAIT  # 总超时时间
        timeout_find = self.settings.TIME_FIND  # 查找超时时间
        start_time = time.time()  # 轮询开始时间
        last_refresh_time = start_time  # 最新刷新时间
        refresh_interval = self.settings.REFRESH_INTERVAL  # 刷新间隔时间

        while time.time() - start_time < timeout:
            try:
                if real_action == "value":
                    val = self._get_element_value(element_name, timeout_find)
                elif real_action == "text":
                    val = self._get_text(element_name, timeout_find)
                else:
                    val = self._get_element_value(element_name, timeout_find)

                if val.strip() == str(expected_value).strip():
                    logger.info(f"值匹配成功: {expected_value}")
                    return
            except PlaywrightTimeoutError:
                pass  # 元素不存在，刷新后继续轮询
            if refresh_interval and time.time() - last_refresh_time >= refresh_interval:
                logger.info(f"元素：{element_name}非期望值{expected_value}，刷新页面继续等待...")
                self.page.reload()
                last_refresh_time = time.time()
            time.sleep(1)  # 轮询节奏

        error_msg = f"等待超时: {element_name} 期望={expected_value}"
        self.take_screenshot(f"等待值超时-{element_name}")
        raise Exception(error_msg)

    @allure.step("等待元素{element_name}出现")
    def wait_for_element_appear(self, element_name):
        total_wait = self.settings.WAIT_ELEMENT_APPEAR
        refresh_interval = self.settings.REFRESH_TIME
        timeout_find = self.settings.TIME_FIND
        start_time = time.time()
        last_refresh_time = start_time

        while time.time() - start_time < total_wait:
            if self.is_element_present(element_name, timeout=timeout_find):
                logger.info(f"元素已出现: {element_name}")
                return True

            if time.time() - last_refresh_time >= refresh_interval:
                logger.info("刷新页面继续等待...")
                self.page.reload()
                last_refresh_time = time.time()

            time.sleep(2)  # 轮询节奏

        error_msg = f"等待元素出现超时: {element_name}"
        self.take_screenshot(f"等待元素超时-{element_name}")
        raise Exception(error_msg)

    @allure.step("执行 SQL 验证")
    def verify_mysql_data(self, sql: str, expected: str):
        logger.info(f"执行 SQL: {sql}")
        result = self._db_utils.execute_query(sql)
        allure.attach(str(result), "SQL 查询结果", allure.attachment_type.TEXT)

        if not self.parse_and_verify_expected(result, expected):
            raise AssertionError(f"数据库验证失败，期望结果: {expected}")
        logger.info("数据库验证通过")

    def parse_and_verify_expected(self, result, expected):
        if expected.lower() in ["empty", "[]", "null", "none"]:
            return len(result) == 0
        if expected.startswith("count:"):
            return len(result) == int(expected.split(":")[1])
        if expected.startswith("count>:"):
            return len(result) > int(expected.split(":")[1])
        if expected.startswith("contains:"):
            v = expected.split(":", 1)[1]
            return any(v in str(val) for row in result for val in row.values())
        if "=" in expected:
            return self.verify_field_values(result, expected)
        if len(result) == 1 and len(result[0]) == 1:
            return str(list(result[0].values())[0]) == expected
        return len(result) > 0

    @staticmethod
    def verify_field_values(result, expected):
        expect_dict = dict(kv.split("=", 1) for kv in expected.split(","))
        for row in result:
            if all(str(row[k]).strip() == v.strip() for k, v in expect_dict.items()):
                return True
        return False

    @allure.step("执行数据库更新 SQL")
    def execute_mysql_update(self, sql: str):
        rows = self._db_utils.execute_update(sql)
        logger.info(f"更新完成，影响行数: {rows}")
        return rows

    def wait_if_loading(self, timeout=None):
        """点击/登录后：若页面存在可见的加载动画，等待其消失；
        依赖配置：settings.LOADING_SELECTOR（当前项目的动画选择器，如 "#loader-wrapper"）
        原理：
          - 未配置 LOADING_SELECTOR → 直接返回
          - loader 不存在或不可见 → 即时返回（count/is_visible 都是同步查询，不等待）
          - loader 可见（动画中）→ 等它 hidden（被移除/隐藏）
        """
        if timeout is None:
            timeout = self.wait_timeout
        # 当前项目未配置加载动画选择器 → 不检测
        if not self.settings.LOADING_SELECTOR:
            logger.info(f"页面未配置加载项,直接跳过")
            return
        loc = self.page.locator(self.settings.LOADING_SELECTOR).first
        # 即时判断：元素不存在 或 不可见（display:none/移除）→ 没有动画，直接返回
        if loc.count() == 0 or not loc.is_visible():
            logger.info(f"页面加载项元素不存在或不可见,加载已完成")
            sleep(3)  # 额外等待3s后开始执行测试
            return
        # 有可见的加载动画 → 等它消失（元素被移除/隐藏）
        try:
            loc.wait_for(state="hidden", timeout=timeout)
            logger.info(f"加载动画已结束: {self.settings.LOADING_SELECTOR}")
        except TimeoutError:
            logger.warning(f"加载动画未在{timeout / 1000}秒内消失: {self.settings.LOADING_SELECTOR}")
